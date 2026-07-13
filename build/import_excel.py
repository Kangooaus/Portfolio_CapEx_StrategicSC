"""
import_excel.py — sync hand-edits made in the Excel workbook back into
capex_portfolio.db.

This is the "keep authoring in Excel" input path: someone opens
downloads/CapEx_StrategicSC_Portfolio.xlsx, edits an input cell (unit cost,
lead time, a program's IRR estimate, an assumption value...) directly in
Excel, saves it, then runs this script to pull that edit into the database.
It is a one-way import (Excel -> DB) triggered explicitly by running the
script -- there is no live sync, so two people editing the same row in
Excel and DB Browser at the same time can still conflict. Re-running
build_workbook.py after an import regenerates the workbook from the DB,
closing the loop.

Only INPUT cells are read -- the same cells the sheet generators marked
BLUE_INPUT when they wrote them. Formula cells (totals, calculated risk
scores, cross-sheet links) are never read back; the database is the only
place those get recomputed. Two fields on S33 are Excel-only presentation
text (Strategic Capacity Impact, Notes) and two more (PRG-002's FX Reserve
and NRE, which are baked into a formula literal rather than their own
cells) are intentionally not round-tripped -- edit those in the database
directly if they change.

Usage:
    python3 import_excel.py                  # dry-run: report the diff, write nothing
    python3 import_excel.py --apply           # apply additions/updates to the DB
    python3 import_excel.py --apply --delete-missing
                                               # also delete DB rows whose natural
                                               # key no longer appears in Excel
    python3 import_excel.py --file path.xlsx  # import from a different workbook copy
"""
import argparse
import sys
from pathlib import Path

import openpyxl

import db

DEFAULT_XLSX = Path(__file__).resolve().parent.parent / "downloads" / "CapEx_StrategicSC_Portfolio.xlsx"
PROGRAM_REFS = ("PRG-001", "PRG-002")


def _program_id_rows(ws, program_id_col):
    """Yield (row_cells,) for every row where program_id_col holds a known
    program ref. Totals/section-header rows never populate that column, so
    this reliably selects only real data rows regardless of sheet layout."""
    for row in ws.iter_rows(min_row=1):
        cell = row[program_id_col - 1]
        if cell.value in PROGRAM_REFS:
            yield row


def _yn(value):
    """Normalize any Yes/No spelling (Excel uses 'Yes' on S03, 'YES' on S16)
    to the DB's canonical 'Yes'/'No' text, case-insensitively -- so a stray
    lowercase 'yes' typed into a cell doesn't silently import as 'No'."""
    return "Yes" if str(value).strip().lower() == "yes" else "No"


# ------------------------------------------------------------- equipment

def read_equipment(wb):
    ws = wb["S02_Equipment_Portfolio"]
    counters = {ref: 0 for ref in PROGRAM_REFS}
    out = []
    for row in _program_id_rows(ws, 19):
        program_ref = row[18].value
        out.append({
            "program_ref": program_ref,
            "equip_id": row[0].value,
            "category": row[1].value,
            "tool_type": row[2].value,
            "supplier": row[3].value,
            "region": row[4].value,
            "currency": row[5].value,
            "unit_cost": row[6].value,
            "qty": row[7].value,
            "install_cost_unit": row[9].value,
            "life_yrs": row[11].value,
            "energy_kwh_hr": row[12].value,
            "maint_rate_pct": row[13].value,
            "throughput_uhr": row[14].value,
            "op_hrs_yr": row[15].value,
            "notes": row[17].value,
            "sort_order": counters[program_ref],
        })
        counters[program_ref] += 1
    return out


# ------------------------------------------------------------- suppliers

def read_suppliers(wb):
    ws = wb["S03_Supplier_Dataset"]
    counters = {ref: 0 for ref in PROGRAM_REFS}
    out = []
    for row in _program_id_rows(ws, 15):
        program_ref = row[14].value
        out.append({
            "program_ref": program_ref,
            "supplier_name": row[0].value,
            "component_type": row[1].value,
            "equip_ref": row[2].value,
            "region": row[3].value,
            "currency": row[4].value,
            "std_lt_wks": row[5].value,
            "current_lt_wks": row[6].value,
            "reliability": row[8].value,
            "single_source": _yn(row[9].value),
            "alt_supplier_available": _yn(row[10].value),
            "notes": row[13].value,
            "sort_order": counters[program_ref],
        })
        counters[program_ref] += 1
    return out


# ------------------------------------------------------------ risk_items

def read_risk_items(wb):
    ws = wb["S16_Supply_Chain_Risk"]
    counters = {ref: 0 for ref in PROGRAM_REFS}
    out = []
    for row in _program_id_rows(ws, 15):
        program_ref = row[14].value
        std_lead = row[3].value
        current_lead = row[4].value
        out.append({
            "program_ref": program_ref,
            "equip_ref": row[0].value,
            "supplier": row[1].value,
            "component": row[2].value,
            "std_lead": None if std_lead in (None, "N/A") else std_lead,
            "current_lead": None if current_lead in (None, "N/A") else current_lead,
            "delay_prob": row[5].value,
            "sched_impact": row[6].value,
            "replace_difficulty": row[7].value,
            "single_source": _yn(row[10].value),
            "buffer_stock": _yn(row[11].value),
            "mitigation": row[12].value,
            "status": row[13].value,
            "sort_order": counters[program_ref],
        })
        counters[program_ref] += 1
    return out


# ----------------------------------------------------------- assumptions

def read_assumptions(wb):
    ws = wb["S01_Disclaimer_Assumptions"]
    known = {a["parameter"]: a for a in db.list_assumptions()}
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=1)):
        param = row[0].value
        if param not in known:
            continue
        value = row[1].value
        section = known[param]["section"]
        entry = {
            "program_ref": None,
            "section": section,
            "parameter": param,
            "unit": row[2].value,
            "source": row[3].value,
            "date_validated": row[4].value,
            "notes": row[5].value,
            "sort_order": known[param]["sort_order"],
        }
        if isinstance(value, str):
            entry["value"] = None
            entry["value_text"] = value
        else:
            entry["value"] = value
            entry["value_text"] = None
        out.append(entry)
    return out


# -------------------------------------------------------------- programs

def read_programs(wb):
    ws = wb["S33_Portfolio_Register"]
    out = []
    for row in ws.iter_rows(min_row=1):
        ref = row[0].value
        if ref not in PROGRAM_REFS:
            continue
        entry = {
            "program_ref": ref,
            "name": row[1].value,
            "sponsor": row[2].value,
            "program_manager": row[3].value,
            "business_unit": row[4].value,
            "phase": row[5].value,
            "status": row[6].value,
            "payback_yrs": row[12].value,
        }
        if ref == "PRG-002":
            # PRG-001's IRR/NPV cells are cross-sheet formulas (computed,
            # not input); PRG-002's are literal BLUE_INPUT values.
            entry["irr"] = row[10].value
            entry["npv"] = row[11].value
        out.append(entry)
    return out


# ------------------------------------------------------------------ diff

_BUSINESS_KEYS = {
    "equipment": [k for k in [
        "category", "tool_type", "supplier", "region", "currency", "unit_cost", "qty",
        "install_cost_unit", "life_yrs", "energy_kwh_hr", "maint_rate_pct",
        "throughput_uhr", "op_hrs_yr", "notes"]],
    "suppliers": ["supplier_name", "component_type", "region", "currency", "std_lt_wks",
                  "current_lt_wks", "reliability", "single_source", "alt_supplier_available", "notes"],
    "risk_items": ["supplier", "std_lead", "current_lead", "delay_prob", "sched_impact",
                   "replace_difficulty", "single_source", "buffer_stock", "mitigation", "status"],
    "assumptions": ["section", "value", "value_text", "unit", "source", "date_validated", "notes"],
    "programs": ["name", "sponsor", "program_manager", "business_unit", "phase", "status",
                 "payback_yrs", "irr", "npv"],
}


def diff_entity(entity, excel_rows, db_rows, key_fn):
    db_by_key = {key_fn(r): r for r in db_rows}
    seen_keys = set()
    changes = []
    for erow in excel_rows:
        key = key_fn(erow)
        seen_keys.add(key)
        old = db_by_key.get(key)
        fields = _BUSINESS_KEYS[entity]
        if old is None:
            changes.append({"key": key, "status": "new", "row": erow, "diff": None})
        else:
            # Fields absent from erow weren't read from Excel (e.g. PRG-001's
            # formula-derived IRR/NPV) -- leave those untouched, not diffed.
            diffs = {f: (old.get(f), erow.get(f)) for f in fields if f in erow and old.get(f) != erow.get(f)}
            if diffs:
                changes.append({"key": key, "status": "changed", "row": erow, "diff": diffs})
    missing = [r for r in db_rows if key_fn(r) not in seen_keys]
    return changes, missing


def _fmt_key(key):
    return "/".join(str(k) for k in key)


def report(entity, changes, missing, key_fn):
    if not changes and not missing:
        return
    print(f"\n{entity}:")
    for c in changes:
        label = _fmt_key(c["key"])
        if c["status"] == "new":
            print(f"  + NEW      {label}")
        else:
            print(f"  ~ CHANGED  {label}")
            for f, (old, new) in c["diff"].items():
                print(f"        {f}: {old!r} -> {new!r}")
    for m in missing:
        print(f"  - MISSING  {_fmt_key(key_fn(m))}  (in DB, not in Excel)")


def run(xlsx_path, apply, delete_missing):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    equipment = read_equipment(wb)
    suppliers = read_suppliers(wb)
    risk_items = read_risk_items(wb)
    assumptions = read_assumptions(wb)
    programs = read_programs(wb)

    eq_key = lambda r: (r["program_ref"], r["equip_id"])
    sup_key = lambda r: (r["program_ref"], r["equip_ref"])
    risk_key = lambda r: (r["program_ref"], r["equip_ref"], r["component"])
    assum_key = lambda r: (r["parameter"],)
    prog_key = lambda r: (r["program_ref"],)

    eq_changes, eq_missing = diff_entity("equipment", equipment, db.list_equipment(), eq_key)
    sup_changes, sup_missing = diff_entity("suppliers", suppliers, db.list_suppliers(), sup_key)
    risk_changes, risk_missing = diff_entity("risk_items", risk_items, db.list_risk_items(), risk_key)
    assum_changes, assum_missing = diff_entity("assumptions", assumptions, db.list_assumptions(), assum_key)
    prog_changes, prog_missing = diff_entity("programs", programs, db.list_programs(), prog_key)

    report("equipment", eq_changes, eq_missing, eq_key)
    report("suppliers", sup_changes, sup_missing, sup_key)
    report("risk_items", risk_changes, risk_missing, risk_key)
    report("assumptions", assum_changes, assum_missing, assum_key)
    report("programs", prog_changes, prog_missing, prog_key)

    total_changes = sum(len(c) for c in (eq_changes, sup_changes, risk_changes, assum_changes, prog_changes))
    total_missing = sum(len(m) for m in (eq_missing, sup_missing, risk_missing, assum_missing, prog_missing))

    if total_changes == 0 and total_missing == 0:
        print("No differences -- database already matches the workbook.")
        return

    print(f"\n{total_changes} row(s) to add/update, {total_missing} row(s) in DB but not in Excel.")

    if not apply:
        print("Dry run -- no changes written. Re-run with --apply to write them.")
        return

    conn = db.get_connection()
    for c in eq_changes:
        db.upsert_equipment(c["row"], conn=conn)
    for c in sup_changes:
        db.upsert_supplier_by_key(c["row"], conn=conn)
    for c in risk_changes:
        db.upsert_risk_item_by_key(c["row"], conn=conn)
    for c in assum_changes:
        row = dict(c["row"])
        existing = next((a for a in db.list_assumptions() if a["parameter"] == row["parameter"]), None)
        if existing:
            row["id"] = existing["id"]
        db.upsert_assumption(row, conn=conn)
    for c in prog_changes:
        merged = dict(db.get_program(c["row"]["program_ref"]) or {}, **c["row"])
        db.upsert_program(merged, conn=conn)

    if delete_missing:
        for m in eq_missing:
            db.delete_equipment(m["program_ref"], m["equip_id"], conn=conn)
        for m in sup_missing:
            db.delete_supplier(m["id"], conn=conn)
        for m in risk_missing:
            db.delete_risk_item(m["id"], conn=conn)
        print(f"Deleted {len(eq_missing) + len(sup_missing) + len(risk_missing)} row(s) missing from Excel.")
    elif total_missing:
        print("(Rows missing from Excel were left in place. Re-run with --delete-missing to remove them.)")

    conn.close()
    print("Applied.")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--file", default=str(DEFAULT_XLSX), help="Path to the workbook to import from")
    ap.add_argument("--apply", action="store_true", help="Write changes to the database (default: dry-run)")
    ap.add_argument("--delete-missing", action="store_true",
                     help="Also delete DB rows whose key no longer appears in Excel (only with --apply)")
    args = ap.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    run(xlsx_path, apply=args.apply, delete_missing=args.delete_missing)


if __name__ == "__main__":
    main()
