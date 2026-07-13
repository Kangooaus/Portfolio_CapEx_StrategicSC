from gen_common import *
from openpyxl.worksheet.hyperlink import Hyperlink

ZONE_LEGEND = [
    (0, "Gray", "Navigation & Governance", "Index, Disclaimer & Assumptions, version control"),
    (1, "Blue", "Program Foundation", "Equipment portfolio, supplier data, should-cost, TCO"),
    (2, "Green", "Planning & Scheduling", "WBS, CPM, Milestone Roadmap, deployment, capacity, demand forecast"),
    (3, "Orange", "Financial & Risk", "AR, cash flow, FX, scenario analysis, supply chain risk"),
    (4, "Purple", "Strategic Supply Chain", "Market analysis, network, sourcing, dual-source, negotiation, inventory, assets"),
    (5, "Red", "Execution & Control", "Gates, ECO, VDC, yield metrics, ROI tracker, scorecard, change tracker"),
    (6, "Teal", "Portfolio & Capital Allocation", "Multi-program register, portfolio dashboard, prioritization scorecard"),
]

DIRECTORY = [
    ("S00", "S00_Index", 0, "New", "Workbook navigation, zone guide, sheet directory", "This page"),
    ("S01", "S01_Disclaimer_Assumptions", 0, "Revised", "Portfolio disclaimer, global model assumptions, version control", "WACC=8%, Uptime=92%, FX rates, version log"),
    ("S02", "S02_Equipment_Portfolio", 1, "Revised", "Master equipment dataset for 2 programs (PRG-001, PRG-002) — costs, energy, throughput", "$13.87M + $2.38M CapEx across 21 platforms"),
    ("S03", "S03_Supplier_Dataset", 1, "Revised", "Supplier lead times, reliability scores, risk tiers across both programs", "21 OEMs; Program ID column for portfolio rollups"),
    ("S04", "S04_Should_Cost_Model", 1, "Preserved", "First-principles should-cost by material, labour, OH, margin", "Should-cost vs quote variance analysis"),
    ("S05", "S05_TCO_Downtime_Model", 1, "Preserved", "10-year TCO, energy, maintenance, downtime cost comparison", "NPV-based lifecycle cost comparison"),
    ("S06", "S06_WBS", 2, "New", "Work Breakdown Structure — all program workstreams and tasks", "8 workstreams, 60+ work packages"),
    ("S07", "S07_CPM_Master_Schedule", 2, "New", "Critical Path Method schedule — activities, float, predecessors", "Critical path = 46 weeks; 39 activities"),
    ("S08", "S08_Program_Milestone_Roadmap", 2, "New", "Visual Gantt-style program roadmap — recruiter-friendly view", "G0 to G5 gates, Apr-2025 to Dec-2025"),
    ("S09", "S09_Deployment_Timeline", 2, "Revised", "Equipment-level deployment schedule by workstream", "46-week program; 6 critical-path items"),
    ("S10", "S10_Capacity_Planning", 2, "Revised", "Equipment capacity model — throughput, utilization, bottleneck ID", "1,040K units/yr capacity; 72.1% utilization"),
    ("S11", "S11_Demand_Forecast", 2, "New", "Time-series demand forecast with scenario banding and accuracy tracking", "750K baseline; 2023-2025 ramp model"),
    ("S12", "S12_AR_Summary", 3, "Revised", "Formal capital appropriations request — Phase II budget & approval", "$16.062M AR; 14.2% IRR; 4.2yr payback"),
    ("S13", "S13_Scenario_Analysis", 3, "Revised", "Scenario A/B/C comparison — CapEx, OpEx, NPV tradeoffs", "Scenario B recommended; $40.04M benchmark"),
    ("S14", "S14_CashFlow_Model", 3, "New", "Merged monthly + quarterly cash flow with S-curve profile", "$14.42M program spend; compressed to 2025"),
    ("S15", "S15_FX_Exposure_Model", 3, "Revised", "EUR/JPY FX risk — baseline, adverse, favorable scenarios", "9/10 line items hedged"),
    ("S16", "S16_Supply_Chain_Risk", 3, "Revised", "Supply chain risk heatmap — 25 components across both programs, 4 risk tiers", "7 CRITICAL; 4 HIGH; single-source analysis"),
    ("S17", "S17_Supply_Market_Analysis", 4, "New", "Supply market structure — concentration, players, pricing trends", "8 commodities; HHI analysis; price trends"),
    ("S18", "S18_Supply_Chain_Network", 4, "New", "Logistics network map — supplier geography, routes, Tier 2 visibility", "15 suppliers; transit times; customs data"),
    ("S19", "S19_Strategic_Sourcing_Pipeline", 4, "New", "Full sourcing workflow — RFI to award, contract status, savings", "12 categories tracked"),
    ("S20", "S20_Dual_MultiSource_Strategy", 4, "New", "Dual/multi-source program — qualification status, risk reduction", "8 commodities; 3 active qual programs"),
    ("S21", "S21_Negotiation_Tracker", 4, "New", "Commercial negotiation log — targets, BATNA, outcomes, savings", "10 negotiations; $537K confirmed savings"),
    ("S22", "S22_Capacity_Supplier_Facing", 4, "New", "Supplier capacity assessment — program share, flex, MOQ, buffer", "15 suppliers; capacity constraint analysis"),
    ("S23", "S23_Inventory_Buffer_Planning", 4, "New", "Inventory strategy — raw materials, WIP, spares, safety stock", "$124.2K spare parts inventory value"),
    ("S24", "S24_Asset_Tracking", 4, "New", "Physical asset register — location, status, lifecycle, warranty", "45 assets; installation dates; replacement plan"),
    ("S25", "S25_Milestone_Gate_Tracker", 5, "Revised", "Phase-gate program tracker — G0 to G5 gates, criteria, blockers", "G0+G1 approved; G2-G5 in progress"),
    ("S26", "S26_Engineering_Change_Orders", 5, "New", "Formal ECO register — origin, scope, cost/schedule impact, approval", "10 ECOs; $107.2K cost impact"),
    ("S27", "S27_Vendor_Design_Changes", 5, "New", "Supplier-initiated design change log — disposition, requalification", "7 VDCs; risk-assessed; approval tracked"),
    ("S28", "S28_Yield_Production_Metrics", 5, "New", "OEE, yield, defect tracking by process tool and period", "96% target yield; OEE model by equipment"),
    ("S29", "S29_Equipment_ROI_Tracker", 5, "Revised", "Post-commissioning actual vs plan — throughput, uptime, revenue impact", "30/90/180-day checkpoints; variance analysis"),
    ("S30", "S30_Executive_Dashboard", 0, "Revised", "Executive KPI summary — all zones linked, recruiter first view", "All key metrics; zone navigation prompts"),
    ("S31", "S31_Vendor_Scorecard", 5, "New", "Weighted vendor scorecard — delivery, quality, commercial, technical", "18 suppliers scored; 4 Preferred, 9 Approved"),
    ("S32", "S32_Change_Tracker", 5, "New", "Engineering change management log — cost, schedule, risk", "12 CRs; $210.5K cost impact; 5.8% contingency used"),
    ("S33", "S33_Portfolio_Register", 6, "New", "Multi-program register — one row per program, live-linked financials & risk", "2 programs; PRG-001 Greenfield, PRG-002 Riverside"),
    ("S34", "S34_Portfolio_Dashboard", 6, "New", "Portfolio-wide KPI rollup across all active/proposed programs", "Total portfolio CapEx, blended risk, program comparison"),
    ("S35", "S35_Portfolio_Prioritization", 6, "New", "Weighted capital allocation scorecard — ranks programs for sequencing", "Greenfield #1 (4.25/5.0), Riverside #2 (3.15/5.0)"),
]


def build(wb):
    ws = wb.create_sheet("S00_Index", 0)
    set_tab_color(ws, 0)
    row = title_block(ws, 0, "S00", "WORKBOOK INDEX — Greenfield Expansion Phase II · CapEx & Supply Chain Portfolio Model",
                       "Prepared by: Sourabh Tarodekar | Contact: sourabh232@gmail.com | "
                       "Purpose: Portfolio demonstration of Engineering Program Management – CapEx & Supply Chain Operations",
                       n_cols=6)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value=(
        "This workbook contains synthetic sample data created for portfolio demonstration purposes only. "
        "The data does not represent proprietary or confidential information from any employer or client."
    )).font = Font(italic=True, size=9, color="808080")
    ws.cell(row=row, column=1).alignment = LEFT_WRAP
    row += 2

    row = section_header(ws, row, 1, "ZONE LEGEND — TAB COLOR GUIDE", 4, 0)
    row = write_headers(ws, row, 1, ["Zone", "Color", "Category", "Purpose"], 0)
    for zone, color, cat, purpose in ZONE_LEGEND:
        ws.cell(row=row, column=1, value=f"Zone {zone}")
        cc = ws.cell(row=row, column=2, value=color)
        cc.fill = zone_fill(zone)
        cc.font = Font(color="FFFFFF", bold=True) if zone != 0 else Font(color="FFFFFF", bold=True)
        ws.cell(row=row, column=3, value=cat)
        nc = ws.cell(row=row, column=4, value=purpose); nc.alignment = LEFT_WRAP
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER_ALL
        row += 1
    row += 1

    row = section_header(ws, row, 1, "SHEET DIRECTORY", 6, 0)
    row = write_headers(ws, row, 1, ["Sheet #", "Sheet Name (click to navigate)", "Zone", "Status",
                                      "What This Sheet Demonstrates", "Key Metrics / Outputs"], 0)
    for sid, name, zone, status, desc, metrics in DIRECTORY:
        ws.cell(row=row, column=1, value=sid)
        link_c = ws.cell(row=row, column=2, value=name)
        link_c.hyperlink = f"#'{name}'!A1"
        link_c.font = Font(color="0000FF", underline="single")
        zc = ws.cell(row=row, column=3, value=f"Zone {zone}")
        zc.fill = zone_fill_light(zone)
        ws.cell(row=row, column=4, value=status)
        dc = ws.cell(row=row, column=5, value=desc); dc.alignment = LEFT_WRAP
        mc = ws.cell(row=row, column=6, value=metrics); mc.alignment = LEFT_WRAP
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER_ALL
        row += 1

    autosize(ws, {1: 8, 2: 32, 3: 9, 4: 11, 5: 52, 6: 40})
    freeze_below(ws, 1)
    return {"sheet": ws.title}
