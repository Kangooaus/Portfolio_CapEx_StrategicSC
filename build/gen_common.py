"""
Shared styling helpers for the CapEx Strategic SC Portfolio workbook rebuild.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

# Zone tab colors (hex, no '#')
ZONE_COLORS = {
    0: "808080",  # Gray - Navigation & Governance
    1: "1F4E78",  # Blue - Program Foundation
    2: "2E7D32",  # Green - Planning & Scheduling
    3: "E67E22",  # Orange - Financial & Risk
    4: "6A1B9A",  # Purple - Strategic Supply Chain
    5: "C0392B",  # Red - Execution & Control
}

ZONE_FILL_LIGHT = {
    0: "D9D9D9",
    1: "D9E6F2",
    2: "DCEEDC",
    3: "FCE4CC",
    4: "E6D9F0",
    5: "F5D6D2",
}

RAG_GREEN = "C6EFCE"
RAG_GREEN_FONT = "006100"
RAG_AMBER = "FFEB9C"
RAG_AMBER_FONT = "9C6500"
RAG_RED = "FFC7CE"
RAG_RED_FONT = "9C0006"

WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=16)
SUBTITLE_FONT = Font(color="404040", italic=True, name="Calibri", size=10)
SECTION_FONT = Font(bold=True, name="Calibri", size=12, color="FFFFFF")
BOLD = Font(bold=True)
BLUE_INPUT = Font(color="0000FF")
GREEN_LINK = Font(color="008000")

THIN = Side(style="thin", color="B7B7B7")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

CUR0 = '#,##0'
CUR2 = '#,##0.00'
USD0 = '"$"#,##0'
USD2 = '"$"#,##0.00'
PCT1 = '0.0%'
PCT2 = '0.00%'


def zone_fill(zone):
    return PatternFill("solid", fgColor=ZONE_COLORS[zone])


def zone_fill_light(zone):
    return PatternFill("solid", fgColor=ZONE_FILL_LIGHT[zone])


def set_tab_color(ws, zone):
    ws.sheet_properties.tabColor = ZONE_COLORS[zone]


def title_block(ws, zone, sheet_code, title, subtitle, n_cols=12, row=1):
    """Writes a title/subtitle banner across n_cols starting at `row`."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=f"{sheet_code} — {title}")
    c.font = TITLE_FONT
    c.fill = zone_fill(zone)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26
    if subtitle:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=n_cols)
        c2 = ws.cell(row=row + 1, column=1, value=subtitle)
        c2.font = SUBTITLE_FONT
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[row + 1].height = 15
    return row + 3  # next free row


def section_header(ws, row, col, text, n_cols, zone):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_cols - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = SECTION_FONT
    c.fill = zone_fill(zone)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1


def write_headers(ws, row, col, headers, zone):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col + i, value=h)
        c.font = WHITE_FONT
        c.fill = zone_fill(zone)
        c.alignment = CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 30
    return row + 1


def write_row(ws, row, col, values, number_formats=None, borders=True):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col + i, value=v)
        if borders:
            c.border = BORDER_ALL
        if number_formats and i in number_formats:
            c.number_format = number_formats[i]
    return row + 1


def autosize(ws, widths):
    """widths: dict col_index(1-based) -> width"""
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def rag_fill(tier):
    tier = (tier or "").upper()
    if "CRIT" in tier or "RED" in tier or "HIGH" in tier and "MOD" not in tier:
        return PatternFill("solid", fgColor=RAG_RED)
    if "MOD" in tier or "AMBER" in tier or "WARN" in tier:
        return PatternFill("solid", fgColor=RAG_AMBER)
    if "LOW" in tier or "GREEN" in tier or "OK" in tier:
        return PatternFill("solid", fgColor=RAG_GREEN)
    return None


def freeze_below(ws, row, col_letter="A"):
    ws.freeze_panes = f"{col_letter}{row}"


def cellref(named, key):
    """named: dict from s01_assumptions.build() mapping param name -> (sheet, row, col). Returns 'Sheet'!$X$Y"""
    sheet, row, col = named[key]
    return f"'{sheet}'!${get_column_letter(col)}${row}"
