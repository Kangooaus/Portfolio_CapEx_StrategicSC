from gen_common import *
from openpyxl.styles import Font


def build(wb):
    ws = wb.create_sheet("S01_Disclaimer_Assumptions")
    set_tab_color(ws, 0)
    row = title_block(ws, 0, "S01", "Disclaimer & Global Model Assumptions",
                       "Program: Greenfield Expansion Phase II  |  All hardcoded inputs are in BLUE — update here ONLY. All other sheets link to this sheet.",
                       n_cols=6)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value=(
        "This workbook contains synthetic sample data created for portfolio demonstration purposes only. "
        "The data does not represent proprietary or confidential information from any employer or client. "
        "Prepared by Sourabh Tarodekar for demonstration of Engineering Program Management – CapEx & Supply "
        "Chain Operations capabilities. Contact: sourabh232@gmail.com"
    )).font = Font(italic=True, size=9, color="808080")
    ws.cell(row=row, column=1).alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 30
    row += 2

    headers = ["Parameter", "Value", "Unit", "Source / Basis", "Date Validated", "Notes"]

    sections = [
        ("ENERGY & UTILITIES", [
            ("Energy Price – Electricity", 0.110, "USD/kWh", "EIA Industrial Rate Survey – US Avg 2024", "2025-01",
             "Used in Annual Energy Cost calculations across S02, S05 TCO model"),
            ("Natural Gas Price", 4.800, "USD/MMBtu", "EIA Natural Gas Industrial Price – 2024", "2025-01",
             "Used for high-temp furnace operating cost estimation"),
            ("Cooling Water Cost", 0.018, "USD/gallon", "Site Utility Rate Schedule Rev C", "2025-01",
             "Applicable to chiller and process cooling loads"),
            ("N2 Bulk Gas Cost", 0.045, "USD/scf", "Site Utility Rate Schedule Rev C", "2025-01",
             "Applicable to gas handling system operating cost"),
        ]),
        ("FINANCIAL & DISCOUNT ASSUMPTIONS", [
            ("Discount Rate (WACC)", 0.08, "—", "Finance Department; WACC model approved Q4-2024", "2025-01",
             "Used in NPV calculations; 8% reflects blended cost of capital"),
            ("Project Horizon", 10, "Years", "Standard CapEx lifecycle policy", "2025-01",
             "All TCO and NPV models use 10-year horizon"),
            ("Tax Rate (Effective)", 0.21, "—", "US Federal + State blended estimate", "2025-01",
             "Applied in AR payback model for after-tax IRR"),
            ("CapEx Depreciation Method", "MACRS 7-yr", "—", "Finance / Tax Counsel guidance", "2025-01",
             "Per AR Summary financial justification"),
            ("Inflation Rate", 0.025, "—", "Corporate finance planning assumption", "2025-01",
             "Used to escalate Year 2-10 OpEx in TCO model"),
        ]),
        ("LABOR RATES", [
            ("Fabrication Shop Rate", 95.00, "USD/hr", "Supplier benchmarking; 2024 MW industrial rate", "2025-01",
             "Used in Should-Cost labor model (S04)"),
            ("Engineering Labor Rate", 120.00, "USD/hr", "Internal engineering rate card Rev 2024", "2025-01",
             "Used in NRE and program labor estimates"),
            ("Installation Labor Rate", 85.00, "USD/hr", "Facilities contractor rate card", "2025-01",
             "Used in on-site installation cost estimates"),
            ("Field Service Rate (Supplier)", 175.00, "USD/hr", "Supplier FSE rate benchmark", "2025-01",
             "Used in downtime / MTTR cost modelling"),
        ]),
        ("EQUIPMENT PERFORMANCE", [
            ("Equipment Uptime Target", 0.92, "—", "Program engineering specification", "2025-01",
             "92% uptime = 608 downtime hrs/yr basis for MTBF model"),
            ("Planned Maintenance Downtime", 0.04, "—", "Maintenance planning schedule", "2025-01",
             "4% of annual hours allocated to preventive maintenance"),
            ("Process Yield Assumption", 0.96, "—", "Engineering process spec", "2025-01",
             "Used in capacity model to calculate effective throughput"),
            ("Production Value per Hour", 4500.00, "USD/hr", "Finance — fully-loaded production cost model", "2025-01",
             "Downtime Cost = Value/hr × MTTR hrs"),
        ]),
        ("SUPPLIER COST MODEL INPUTS", [
            ("Supplier Margin – Low", 0.15, "—", "Benchmarking; competitive-bid suppliers", "2025-01",
             "Lower bound; applied to multi-source components"),
            ("Supplier Margin – High", 0.25, "—", "Benchmarking; sole-source suppliers", "2025-01",
             "Upper bound; applied to single-source components"),
            ("Material Scrap Factor – Low", 0.05, "—", "Engineering; precision machined parts", "2025-01",
             "5% scrap on raw material purchases"),
            ("Material Scrap Factor – High", 0.08, "—", "Engineering; cast / welded assemblies", "2025-01",
             "8% scrap on complex fabrications"),
            ("Material Price – Steel (304 SS)", 3.800, "USD/kg", "Metal bulletin Q1-2025", "2025-01",
             "Used in vacuum chamber and furnace body cost estimates"),
            ("Material Price – Aluminum", 4.200, "USD/kg", "Metal bulletin Q1-2025", "2025-01",
             "Used in robotics frame and structural components"),
            ("Material Price – Copper", 9.500, "USD/kg", "COMEX spot Q1-2025", "2025-01",
             "Used in transformer winding and electrical bus cost estimates"),
        ]),
        ("DEPLOYMENT TIMELINE ASSUMPTIONS", [
            ("Supplier Manufacturing Duration", 20, "Weeks", "Program schedule baseline Rev 1.0", "2025-01",
             "Range: 16–24 wks depending on equipment complexity"),
            ("Factory Acceptance Testing (FAT)", 2, "Weeks", "Standard FAT protocol – all equipment", "2025-01",
             "2 weeks minimum; may extend for complex systems"),
            ("International Shipping Duration", 3, "Weeks", "Freight forwarder estimate; ocean freight", "2025-01",
             "Air freight contingency +50% cost; used for risk scenario"),
            ("On-Site Installation", 4, "Weeks", "Facilities project plan Rev A", "2025-01",
             "Per equipment platform; parallel installation possible"),
            ("Utility Hookup", 2, "Weeks", "Facilities electrical + mechanical schedule", "2025-01",
             "Includes inspection and commissioning sign-off"),
            ("Process Qualification (PQ)", 5, "Weeks", "Process engineering qualification plan", "2025-01",
             "Range: 4–6 wks; 5 wks used as baseline"),
        ]),
        ("FX SENSITIVITY ASSUMPTIONS", [
            ("USD/EUR Exchange Rate (Baseline)", 1.09, "USD per EUR", "Bloomberg rates Q1-2025 average", "2025-01",
             "Baseline conversion rate for EUR-denominated equipment"),
            ("USD/JPY Exchange Rate (Baseline)", 149.5, "JPY per USD", "Bloomberg rates Q1-2025 average", "2025-01",
             "Baseline conversion rate for JPY-denominated equipment"),
            ("EUR Sensitivity – Adverse", 1.15, "USD per EUR", "FX stress scenario +6%", "2025-01",
             "1% move in EUR/USD = ~$85K portfolio cost impact"),
            ("EUR Sensitivity – Favorable", 1.03, "USD per EUR", "FX stress scenario -6%", "2025-01",
             "Favorable scenario used in S13 Scenario Analysis"),
            ("JPY Sensitivity – Adverse", 142.00, "JPY per USD", "FX stress scenario (JPY strengthens)", "2025-01",
             "Stronger JPY = higher USD cost for JPY equipment"),
            ("JPY Sensitivity – Favorable", 157.00, "JPY per USD", "FX stress scenario (JPY weakens)", "2025-01",
             "Weaker JPY = lower USD cost for JPY equipment"),
        ]),
    ]

    # remember key cell locations for named ranges
    named = {}

    for sect_title, rows_data in sections:
        row = section_header(ws, row, 1, sect_title, 6, 0)
        row = write_headers(ws, row, 1, headers, 0)
        for name, val, unit, source, date, note in rows_data:
            fmt = None
            if isinstance(val, float) and val < 1:
                fmt = PCT1
            elif isinstance(val, (int, float)) and unit.startswith("USD"):
                fmt = USD2 if isinstance(val, float) and val != int(val) else USD0
            r0 = row
            ws.cell(row=row, column=1, value=name)
            vc = ws.cell(row=row, column=2, value=val)
            vc.font = Font(color="0000FF")
            if fmt:
                vc.number_format = fmt
            ws.cell(row=row, column=3, value=unit)
            ws.cell(row=row, column=4, value=source)
            ws.cell(row=row, column=5, value=date)
            nc = ws.cell(row=row, column=6, value=note)
            nc.alignment = LEFT_WRAP
            for c in range(1, 7):
                ws.cell(row=r0, column=c).border = BORDER_ALL
            named[name] = (ws.title, row, 2)
            row += 1
        row += 1

    # Version control log
    row = section_header(ws, row, 1, "VERSION CONTROL LOG — Model Revision History & Change Tracking", 7, 0)
    vheaders = ["Version", "Date", "Owner / Author", "Change Description", "Impact on Model", "Approval Status", "Notes / Reference"]
    row = write_headers(ws, row, 1, vheaders, 0)
    versions = [
        ("V1.0", "2025-02-10", "Sourabh Tarodekar (TPM)",
         "Initial workbook created. Equipment portfolio dataset populated with 15 platforms from supplier quotes.",
         "Baseline established. All formulas reference V1.0 assumptions.", "APPROVED",
         "Baseline model — AR-2025-0082 submission basis"),
        ("V1.1", "2025-02-14", "S. Kim (Supply Chain)",
         "Updated Pfeiffer vacuum pump (EQ-002) lead time from 16 to 26 weeks based on supplier confirmation.",
         "Deployment timeline extended. EQ-002 now on critical path. Risk flag updated to CRITICAL.", "APPROVED",
         "Supplier email confirmation 2025-02-13 on file"),
        ("V1.2", "2025-02-18", "R. Patel (Finance)",
         "Discount rate updated from 7.5% to 8.0% per Q1-2025 corporate WACC guidance.",
         "NPV reduced by ~$420K. TCO model and scenario analysis recalculated.", "APPROVED",
         "Finance guidance memo FIN-2025-007"),
        ("V1.3", "2025-02-24", "T. Okonkwo (Facilities)",
         "Added EQ-004 crane hire and floor reinforcement to installation cost — +$62K.",
         "Total installation cost increased. CR-001 raised. Change tracker updated.", "APPROVED",
         "CR-001 approved by PM 2025-02-24"),
        ("V1.4", "2025-03-05", "M. Chen (Process Eng.)",
         "Equipment uptime revised from 90% to 92% following process engineering specification review.",
         "Annual capacity increased. Utilization % reduced. Capacity model updated.", "APPROVED",
         "Process spec document PE-2025-014 Rev A"),
        ("V1.5", "2025-03-10", "R. Patel (Finance)",
         "FX sensitivity assumptions added. EUR/JPY baseline and stress scenarios populated.",
         "FX exposure model now active. EUR/USD baseline 1.09, JPY/USD baseline 149.5.", "APPROVED",
         "Bloomberg rates Q1-2025 average"),
        ("V1.6", "2025-03-17", "S. Kim (Supply Chain)",
         "Supplier risk scores recalculated after Daifuku confirmed 30-week manufacturing lead time for EQ-006.",
         "EQ-006 risk tier escalated to CRITICAL. Supply chain risk index increased.", "APPROVED",
         "Daifuku PO confirmation email"),
        ("V1.7", "2025-03-19", "Sourabh Tarodekar (TPM)",
         "AR Summary populated and submitted to VP Engineering + CFO. Budget: $40.04M (Scenario B benchmark).",
         "Formal program approval process initiated. Model frozen for AR submission.", "APPROVED",
         "AR-2025-0082 submitted to Finance"),
        ("V1.8", "2025-04-05", "R. Patel (Finance)",
         "AR approved. POs placed for long-lead items. Spend tracking model activated.",
         "Program execution phase begins. Cash flow schedule now active.", "APPROVED",
         "CFO approval email 2025-04-05"),
        ("V1.9", "2025-05-12", "P. Singh (IT/OT)",
         "IT/OT DMZ infrastructure cost ($45K) added to scope. CR-009 raised.",
         "Contingency draw of $45K. Change tracker updated. Contingency at 5.8%.", "UNDER REVIEW",
         "CR-009 pending PM approval"),
        ("V2.0", "2025-05-19", "Sourabh Tarodekar (TPM)",
         "Version 2.0 release: Cash flow schedule, FX exposure model, and ROI tracker sheets added.",
         "Workbook extended with 4 new analytical sheets. Full model refresh completed.", "APPROVED",
         "Issued to Finance and Engineering Leadership"),
    ]
    for v in versions:
        r0 = row
        for i, val in enumerate(v):
            c = ws.cell(row=row, column=1 + i, value=val)
            c.border = BORDER_ALL
            if i in (3, 4, 6):
                c.alignment = LEFT_WRAP
        status_cell = ws.cell(row=row, column=6)
        fill = rag_fill("GREEN" if v[5] == "APPROVED" else "AMBER")
        if fill:
            status_cell.fill = fill
        row += 1

    autosize(ws, {1: 30, 2: 14, 3: 14, 4: 42, 5: 14, 6: 46, 7: 30})
    freeze_below(ws, 1)
    return named
