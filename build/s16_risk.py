from gen_common import *
from openpyxl.formatting.rule import CellIsRule

# EquipRef, Supplier, Component, StdLT, CurLT, DelayProb, SchedImpact, ReplaceDiff, SingleSource, Buffer, Mitigation, Status
RISK_ITEMS = [
    ("EQ-001", "Edwards Vacuum", "Dry Screw Pump – Complete Unit", 14, 18, 0.35, 3, 2, "No", "No",
     "Expedite order; confirm delivery schedule with supplier weekly; pre-qualify Busch as alt", "IN PROGRESS"),
    ("EQ-002", "Pfeiffer Vacuum", "Turbomolecular Pump – Complete Unit", 16, 26, 0.55, 5, 5, "YES", "No",
     "CRITICAL PATH: Escalate to VP Supply Chain; evaluate rebuilt unit as bridge; initiate alt vendor qual", "ESCALATED"),
    ("EQ-002", "Pfeiffer Vacuum", "Turbomolecular Controller Electronics", 12, 18, 0.45, 4, 5, "YES", "No",
     "Controller firmware proprietary; no drop-in alt; request safety stock commitment from supplier", "ESCALATED"),
    ("EQ-003", "Thermco Systems", "Tube Furnace – Hot Zone Assembly", 18, 20, 0.20, 4, 3, "No", "No",
     "BTU International validated as alt; maintain contact; include alt option in RFQ", "MONITORED"),
    ("EQ-003", "Thermco Systems", "8-Zone Heating Element Set", 10, 12, 0.15, 3, 2, "No", "Yes",
     "2-set buffer stock maintained; supplier reliability high; no immediate action required", "OK"),
    ("EQ-004", "Tokyo Electron (TEL)", "Batch Furnace – Complete System", 20, 28, 0.40, 5, 4, "No", "No",
     "JPY FX hedge confirmed; increase FAT scope to 3 weeks; dedicated TEL PM assigned", "IN PROGRESS"),
    ("EQ-004", "Tokyo Electron (TEL)", "Process Recipe Transfer Package", None, None, 0.30, 4, 4, "YES", "No",
     "Non-tangible risk: recipe IP held by TEL; require contractual data transfer clause at PO stage", "LEGAL REVIEW"),
    ("EQ-005", "Brooks Automation", "SCARA Robot – Mechanical Assembly", 12, 14, 0.15, 3, 3, "No", "No",
     "Brooks delivery historically reliable; no action required; standard monitoring", "OK"),
    ("EQ-005", "Brooks Automation", "MES Software Integration Module", None, None, 0.45, 4, 4, "YES", "No",
     "Integration complexity risk — not supply risk; assign Brooks SW engineer on-site 4 wks pre-FAT", "IN PROGRESS"),
    ("EQ-006", "Daifuku", "Overhead Track System – Civil Interface", 22, 30, 0.50, 4, 5, "YES", "No",
     "Civil drawings required 8 wks pre-install; freeze layout by Week 6; no alt track supplier", "ESCALATED"),
    ("EQ-007", "Air Liquide Eng.", "Gas Cabinet – Complete Assembly", 16, 22, 0.40, 3, 3, "No", "No",
     "Quote under negotiation; delay increases if PO not placed by target date; expedite sourcing decision", "IN PROGRESS"),
    ("EQ-007", "Air Liquide Eng.", "Hazmat Permit – Regulatory", None, None, 0.35, 5, 5, "YES", "No",
     "Permit lead time 8-12 wks; submit application immediately; failure impacts entire gas system install", "ACTION REQ'D"),
    ("EQ-008", "Linde Engineering", "Bulk Gas Storage Vessel", 24, 32, 0.35, 4, 3, "No", "No",
     "Long-lead vessel; PO must be placed by program Week 4; civil pad work parallel path", "IN PROGRESS"),
    ("EQ-009", "Siemens AG", "S7-1500 PLC Hardware Set", 10, 14, 0.15, 2, 2, "No", "Yes",
     "Strong supply reliability; buffer stock of 1 spare CPU maintained; low risk", "OK"),
    ("EQ-009", "Siemens AG", "Cybersecurity Compliance Review", None, None, 0.25, 3, 3, "YES", "No",
     "Internal InfoSec review adds 3 wks; submit documentation package 6 wks pre-installation", "IN PROGRESS"),
    ("EQ-011", "ABB Ltd", "2 MVA Dry-Type Transformer", 14, 20, 0.25, 4, 3, "No", "No",
     "Reuse of legacy unit partially mitigates; new unit on order; monitor delivery", "MONITORED"),
    ("EQ-012", "Daikin Applied", "Chiller – Complete Package", 16, 20, 0.20, 3, 2, "No", "No",
     "Glycol loop interface design must be finalized prior to FAT; coordinate with mechanical team", "IN PROGRESS"),
    ("EQ-014", "Applied Materials", "RTP Tool – Complete System", 20, 24, 0.25, 5, 4, "No", "No",
     "Highest-value item; AMAT requires 30% deposit at PO; finance pre-approval gate required", "ACTION REQ'D"),
    ("EQ-015", "MiR (Mobile Industrial Robots)", "AMR Fleet – 10 Units", 14, 18, 0.30, 2, 2, "No", "No",
     "Wi-Fi infrastructure survey required 6 wks pre-delivery; coordinate with IT/OT team", "MONITORED"),
]

# Program 2 — Riverside Automation Upgrade Phase I (PRG-002)
RISK_ITEMS_PRG2 = [
    ("RV-001", "Fanuc America", "Palletizing Robot – Complete Unit", 10, 12, 0.15, 2, 2, "No", "No",
     "Standard robot platform; low risk; monitor delivery per standard cadence", "OK"),
    ("RV-002", "Krones AG", "Cartoning System – Complete Unit", 20, 24, 0.30, 4, 4, "YES", "No",
     "Single-source specialized tech; no qualified alt; monitor delivery weekly during build", "IN PROGRESS"),
    ("RV-003", "Fetch Robotics", "AGV Fleet – Complete Units", 8, 9, 0.10, 2, 1, "No", "No",
     "Mature platform; low supply risk; standard monitoring", "OK"),
    ("RV-004", "Cognex Corporation", "Vision Inspection – Complete Units", 6, 7, 0.10, 2, 1, "No", "Yes",
     "Commodity vision hardware; Keyence pre-qualified alt on file", "OK"),
    ("RV-005", "Rockwell Automation", "WCS Integration – Software/Hardware", 8, 9, 0.20, 3, 2, "No", "No",
     "Standard PlantPAx platform; integrator has strong track record", "MONITORED"),
    ("RV-006", "Dematic Corp", "Conveyor & Sortation – Complete System", 16, 20, 0.25, 3, 3, "No", "No",
     "Complex site-specific integration; phased cutover plan mitigates schedule risk", "MONITORED"),
]

HEADERS = ["Equip Ref", "Supplier", "Component / Sub-System", "Std Lead (wks)", "Current Lead (wks)",
           "Delay Prob. (0-1)", "Sched. Impact (1-5)", "Replace Difficulty (1-5)", "Composite Risk Score",
           "Risk Tier", "Single Source?", "Buffer Stock?", "Mitigation Action", "Status", "Program ID"]


def _write_risk_rows(ws, r, rows, program_id):
    for (eq, sup, comp, stdlt, curlt, prob, sched, repl, single, buf, mit, status) in rows:
        ws.cell(row=r, column=1, value=eq)
        ws.cell(row=r, column=2, value=sup)
        ws.cell(row=r, column=3, value=comp)
        ws.cell(row=r, column=4, value=stdlt if stdlt else "N/A").font = BLUE_INPUT
        ws.cell(row=r, column=5, value=curlt if curlt else "N/A").font = BLUE_INPUT
        ws.cell(row=r, column=6, value=prob).number_format = PCT1
        ws.cell(row=r, column=6).font = BLUE_INPUT
        ws.cell(row=r, column=7, value=sched).font = BLUE_INPUT
        ws.cell(row=r, column=8, value=repl).font = BLUE_INPUT
        score = ws.cell(row=r, column=9, value=f"=F{r}*G{r}*H{r}")
        score.number_format = CUR2
        ws.cell(row=r, column=10,
                value=f'=IF(OR(K{r}="YES",I{r}>20),IF(I{r}>20,"CRITICAL",IF(K{r}="YES","CRITICAL","HIGH")),IF(I{r}>8,"HIGH",IF(I{r}>2,"MODERATE","LOW")))')
        ws.cell(row=r, column=11, value=single)
        ws.cell(row=r, column=12, value=buf)
        mc = ws.cell(row=r, column=13, value=mit); mc.alignment = LEFT_WRAP
        ws.cell(row=r, column=14, value=status)
        pid = ws.cell(row=r, column=15, value=program_id)
        pid.font = BLUE_INPUT
        pid.alignment = CENTER
        for c in range(1, 16):
            ws.cell(row=r, column=c).border = BORDER_ALL
        r += 1
    return r


def _summary_block(ws, row, first_row, last_row, title):
    row = section_header(ws, row, 1, title, 2, 3)
    stats = [
        ("Total Components Tracked", f"=COUNTA(A{first_row}:A{last_row})"),
        ("Single-Source Components", f'=COUNTIF(K{first_row}:K{last_row},"YES")'),
        ("Avg Composite Risk Score", f"=AVERAGE(I{first_row}:I{last_row})"),
        ("Max Risk Score (Worst Item)", f"=MAX(I{first_row}:I{last_row})"),
        ("CRITICAL Tier Items", f'=COUNTIF(J{first_row}:J{last_row},"CRITICAL")'),
        ("HIGH Tier Items", f'=COUNTIF(J{first_row}:J{last_row},"HIGH")'),
    ]
    stat_rows = {}
    keys = ["total", "single_source", "avg_risk", "max_risk", "critical_count", "high_count"]
    for key, (label, formula) in zip(keys, stats):
        ws.cell(row=row, column=1, value=label).border = BORDER_ALL
        c = ws.cell(row=row, column=2, value=formula)
        c.number_format = CUR2
        c.font = BOLD
        c.border = BORDER_ALL
        stat_rows[key] = row
        row += 1
    return row, stat_rows


def build(wb):
    ws = wb.create_sheet("S16_Supply_Chain_Risk")
    set_tab_color(ws, 3)
    row = title_block(ws, 3, "S16", "Supply Chain Risk & Lead-Time Analysis — CapEx Equipment Program",
                       "Risk Score = Delay Probability x Schedule Impact Score x Replacement Difficulty | "
                       "Heatmap tiers: LOW / MODERATE / HIGH / CRITICAL | Single-source components auto-flagged CRITICAL if Risk Score > 20 | "
                       "Program ID column enables portfolio-wide rollups",
                       n_cols=15)
    header_row = write_headers(ws, row, 1, HEADERS, 3)
    first_row = header_row
    r = header_row
    r = _write_risk_rows(ws, r, RISK_ITEMS, "PRG-001")
    last_row = r - 1
    row = r + 2

    ws.conditional_formatting.add(f"J{first_row}:J{last_row}",
                                   CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill("solid", fgColor=RAG_RED)))
    ws.conditional_formatting.add(f"J{first_row}:J{last_row}",
                                   CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill("solid", fgColor="FFD9B3")))
    ws.conditional_formatting.add(f"J{first_row}:J{last_row}",
                                   CellIsRule(operator="equal", formula=['"MODERATE"'], fill=PatternFill("solid", fgColor=RAG_AMBER)))
    ws.conditional_formatting.add(f"J{first_row}:J{last_row}",
                                   CellIsRule(operator="equal", formula=['"LOW"'], fill=PatternFill("solid", fgColor=RAG_GREEN)))

    row = section_header(ws, row, 1, "SUPPLY CHAIN RISK HEATMAP SUMMARY — PRG-001 Greenfield Expansion Phase II", 4, 3)
    row = write_headers(ws, row, 1, ["Risk Tier", "Count", "% of Items", "Action Required"], 3)
    for tier, action in [
        ("CRITICAL", "Immediate VP-level escalation; daily status; dedicated mitigation team"),
        ("HIGH", "Weekly executive review; mitigation plan required within 5 business days"),
        ("MODERATE", "Bi-weekly review; mitigation plan in place; monitor for deterioration"),
        ("LOW", "Standard monitoring; quarterly supplier check-in sufficient"),
    ]:
        ws.cell(row=row, column=1, value=tier)
        ws.cell(row=row, column=2, value=f'=COUNTIF(J{first_row}:J{last_row},A{row})')
        pct = ws.cell(row=row, column=3, value=f"=B{row}/{last_row - first_row + 1}"); pct.number_format = PCT1
        ws.cell(row=row, column=4, value=action).alignment = LEFT_WRAP
        f = rag_fill(tier)
        if f:
            ws.cell(row=row, column=1).fill = f
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER_ALL
        row += 1
    row += 1

    row, prg1_stats = _summary_block(ws, row, first_row, last_row, "AGGREGATE RISK STATISTICS — PRG-001 Greenfield Expansion Phase II")
    row += 1

    # --- Program 2 block ---
    row = section_header(ws, row, 1, "PRG-002 RISK REGISTER — Riverside Automation Upgrade Phase I", 15, 3)
    prg2_header_row = write_headers(ws, row, 1, HEADERS, 3)
    prg2_first_row = prg2_header_row
    r = prg2_header_row
    r = _write_risk_rows(ws, r, RISK_ITEMS_PRG2, "PRG-002")
    prg2_last_row = r - 1
    row = r + 1

    ws.conditional_formatting.add(f"J{prg2_first_row}:J{prg2_last_row}",
                                   CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill("solid", fgColor=RAG_RED)))
    ws.conditional_formatting.add(f"J{prg2_first_row}:J{prg2_last_row}",
                                   CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill("solid", fgColor="FFD9B3")))
    ws.conditional_formatting.add(f"J{prg2_first_row}:J{prg2_last_row}",
                                   CellIsRule(operator="equal", formula=['"MODERATE"'], fill=PatternFill("solid", fgColor=RAG_AMBER)))
    ws.conditional_formatting.add(f"J{prg2_first_row}:J{prg2_last_row}",
                                   CellIsRule(operator="equal", formula=['"LOW"'], fill=PatternFill("solid", fgColor=RAG_GREEN)))

    row, prg2_stats = _summary_block(ws, row, prg2_first_row, prg2_last_row, "AGGREGATE RISK STATISTICS — PRG-002 Riverside Automation Upgrade Phase I")

    autosize(ws, {1: 9, 2: 22, 3: 32, 4: 10, 5: 11, 6: 11, 7: 10, 8: 12, 9: 13, 10: 10, 11: 10, 12: 10, 13: 50, 14: 13, 15: 12})
    freeze_below(ws, header_row + 1)
    return {
        "sheet": ws.title, "first_row": first_row, "last_row": last_row,
        "program_id_col": "O",
        "prg2_first_row": prg2_first_row, "prg2_last_row": prg2_last_row,
        "prg1_avg_risk_cell": f"B{prg1_stats['avg_risk']}",
        "prg1_single_source_cell": f"B{prg1_stats['single_source']}",
        "prg2_avg_risk_cell": f"B{prg2_stats['avg_risk']}",
        "prg2_single_source_cell": f"B{prg2_stats['single_source']}",
        "prg2_total_cell": f"B{prg2_stats['total']}",
    }
