from pathlib import Path

import openpyxl
import s00_index, s01_assumptions, s02_equipment, s03_supplier, s04_shouldcost, s05_tco
import s06_wbs, s07_cpm, s08_roadmap, s09_deployment, s10_capacity, s11_demand
import s12_ar_summary, s13_scenario, s14_cashflow, s15_fx, s16_risk
import s17_market, s18_network, s19_sourcing, s20_dualsource, s21_negotiation, s22_capacity_supplier, s23_inventory, s24_assets
import s25_gates, s26_eco, s27_vdc, s28_yield, s29_roi, s31_scorecard, s32_changetracker
import s30_dashboard
import s33_portfolio_register, s34_portfolio_dashboard, s35_portfolio_prioritization

OUT_PATH = Path(__file__).resolve().parent.parent / "downloads" / "CapEx_StrategicSC_Portfolio.xlsx"


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Zone 0 (index built last so it can appear first; placeholder order handled by sheet index)
    named = s01_assumptions.build(wb)

    # Zone 1
    s02r = s02_equipment.build(wb, named)
    s03r = s03_supplier.build(wb)
    s04_shouldcost.build(wb, named)
    s05_tco.build(wb, named)

    # Zone 2
    s06_wbs.build(wb)
    s07r = s07_cpm.build(wb)
    s08_roadmap.build(wb)
    s09r = s09_deployment.build(wb)
    s10r = s10_capacity.build(wb, named)
    s11_demand.build(wb)

    # Zone 3
    s12r = s12_ar_summary.build(wb, s02r, s03r)
    s13_scenario.build(wb)
    s14_cashflow.build(wb, (s12r["sheet"], s12r["total_ar_cell"]))
    s15_fx.build(wb, s02r, named)
    s16r = s16_risk.build(wb)

    # Zone 4
    s17_market.build(wb)
    s18_network.build(wb)
    s19_sourcing.build(wb)
    s20_dualsource.build(wb)
    s21_negotiation.build(wb)
    s22_capacity_supplier.build(wb)
    s23_inventory.build(wb, named)
    s24_assets.build(wb)

    # Zone 5
    s25r = s25_gates.build(wb)
    s26_eco.build(wb)
    s27_vdc.build(wb)
    s28r = s28_yield.build(wb)
    s29_roi.build(wb)
    s31_scorecard.build(wb)
    s32_changetracker.build(wb, (s12r["sheet"], s12r["contingency_cell"]))

    # Zone 0 continued: Dashboard + Index (inserted at front)
    s30_dashboard.build(wb, {
        "s02": s02r, "s03": s03r, "s07": s07r, "s09": s09r, "s10": s10r,
        "s12": s12r, "s16": s16r, "s25": s25r, "s28": s28r, "s29": None,
    })
    s00_index.build(wb)

    # Zone 6: Portfolio & Capital Allocation (multi-program rollup layer)
    s33r = s33_portfolio_register.build(wb, s02r, s03r, s16r, s12r)
    s34_portfolio_dashboard.build(wb, s33r)
    s35_portfolio_prioritization.build(wb, s33r)

    # Reorder sheets: S00 Index -> Portfolio layer (S34/S33/S35) -> Program 1 detail (S01-S32)
    order = ["S00_Index", "S34_Portfolio_Dashboard", "S33_Portfolio_Register", "S35_Portfolio_Prioritization",
             "S01_Disclaimer_Assumptions",
             "S02_Equipment_Portfolio", "S03_Supplier_Dataset", "S04_Should_Cost_Model", "S05_TCO_Downtime_Model",
             "S06_WBS", "S07_CPM_Master_Schedule", "S08_Program_Milestone_Roadmap", "S09_Deployment_Timeline",
             "S10_Capacity_Planning", "S11_Demand_Forecast",
             "S12_AR_Summary", "S13_Scenario_Analysis", "S14_CashFlow_Model", "S15_FX_Exposure_Model", "S16_Supply_Chain_Risk",
             "S17_Supply_Market_Analysis", "S18_Supply_Chain_Network", "S19_Strategic_Sourcing_Pipeline",
             "S20_Dual_MultiSource_Strategy", "S21_Negotiation_Tracker", "S22_Capacity_Supplier_Facing",
             "S23_Inventory_Buffer_Planning", "S24_Asset_Tracking",
             "S25_Milestone_Gate_Tracker", "S26_Engineering_Change_Orders", "S27_Vendor_Design_Changes",
             "S28_Yield_Production_Metrics", "S29_Equipment_ROI_Tracker",
             "S30_Executive_Dashboard", "S31_Vendor_Scorecard", "S32_Change_Tracker"]
    wb._sheets = [wb[name] for name in order]
    wb.active = 0

    wb.save(OUT_PATH)
    print("Saved:", OUT_PATH)
    print("Sheets:", len(wb.sheetnames))
    print(wb.sheetnames)


if __name__ == "__main__":
    main()
